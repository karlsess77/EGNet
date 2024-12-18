import argparse
import os
import matplotlib.pyplot as plt
from matplotlib import gridspec
import math

import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader
import numpy as np
from torchvision import transforms
from tqdm import tqdm
import torch.nn.functional as F
from PIL import Image

from network.utils.tumor_edge_dect import process_labels, gauss_labels
from network.entropy import calculate_entropy
from network.EGNet import SI_Unet

from utils.dataset_processing import MyDataset
from monai import transforms
from monai.losses import DiceCELoss, DiceLoss
from monai.metrics import compute_hausdorff_distance, compute_average_surface_distance, DiceMetric
from openpyxl import Workbook, load_workbook

parser = argparse.ArgumentParser(description="diffusion segmentation pipeline")


parser.add_argument("--in_channels", default=3, type=int, help='base channel of UNet')
parser.add_argument("--out_channels", default=2, type=int, help='out channel of UNet')



def test_model(model, test_loader, excel_path, pth_pathdir):
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model.to(device)
    model.eval()

    with torch.no_grad():
        dice_qian = []
        hausdorff_distances = []
        average_distances = []
        test_num = 0
        for data, target in test_loader:
            
            test_num = test_num + 1
            inputs, labels = data.to(device), target.to(device)
            labels_clone = labels.clone().detach()


            labels_clone = F.one_hot(labels_clone, num_classes=2).permute(0, 3, 1, 2)
            labels = torch.unsqueeze(labels, dim=1).to(device)

            outputs, edge_info = model(inputs, is_training=False)

            outputs_softmax = F.softmax(outputs, dim=1)
            outputs_softmax = torch.where(outputs_softmax == torch.max(outputs_softmax, dim=1, keepdim=True).values, 1, 0)
            show_output_softmax = outputs_softmax.cpu()

            dice_loss = DiceMetric(include_background=False)(outputs_softmax, labels_clone)
            dice_metric = dice_loss.mean().detach().cpu().numpy().item()

            hausdorff_dist = compute_hausdorff_distance(outputs_softmax, labels_clone,
                                                    distance_metric="euclidean", percentile=95).cpu()
            average_dist = compute_average_surface_distance(outputs_softmax, labels_clone).cpu()

            workbook = load_workbook(excel_path)
            sheet = workbook.active
            sheet.append([test_num, dice_metric, hausdorff_dist.item(), average_dist.item()])
            workbook.save(excel_path)

            if not (math.isnan(hausdorff_dist) or math.isinf(average_dist)):
                dice_qian.append(dice_metric)

                hausdorff_distances.append(hausdorff_dist.item())

                average_distances.append(average_dist.item())

        average_dice = np.mean(dice_qian)
        average_huasdorff = np.mean(hausdorff_distances)
        average_average_distances = np.mean(average_distances)
        print(f'Average Dice Score on Test Set: {average_dice}, Average Hausdorff Distances on Test Set: {average_huasdorff}, Average Average Distances on Test Set: {average_average_distances},')

def main():
    args = parser.parse_args()
    pth_pathdir = (f'lambda_0.8beta_1.0_bestwish')
    test(args, pth_pathdir)

def test(args, pth_pathdir):
    excel_path = f'./outputs/bestresult/metrics.xlsx'

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    if os.path.exists(excel_path):
        workbook = load_workbook(excel_path)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Metrics'
        sheet.append(["Test Number", "Dice Score", "Hausdorff Distance", "Average Surface Distance"])
        workbook.save(excel_path)

    # Load test dataset
    test_data = MyDataset('test')
    test_loader = DataLoader(test_data, batch_size=1, num_workers=0, pin_memory=True, shuffle=False)

    # Initialize and load trained model
    model = SI_Unet()

    model_size = 0
    for param in model.parameters():
        model_size += param.data.nelement()
    print('Model params: %.2f M' % (model_size / 1024 / 1024))

    path = f'./rectal_tumor_seg/best_{pth_pathdir}.pth'
    print(path)
    checkpoint = torch.load(path)
    model_state_dict = model.state_dict()
    new_state_dict = {}
    for k, v in checkpoint.items():
        if k in model_state_dict and v.size() == model_state_dict[k].size():
            new_state_dict[k] = v

    model.load_state_dict(new_state_dict, strict=False)

    # Test the model
    test_model(model, test_loader, excel_path, pth_pathdir)

if __name__ == '__main__':
    main()